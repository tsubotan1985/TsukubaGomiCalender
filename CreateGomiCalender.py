import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import openpyxl
from icalendar import Calendar, Event
from datetime import datetime, time, timedelta
import pytz
import uuid

class GomiCalendarApp:
    def __init__(self, root):
        self.root = root
        self.root.title("ゴミ収集カレンダー作成")
        self.root.geometry("600x400")

        # タイムゾーンの設定
        self.timezone = pytz.timezone('Asia/Tokyo')

        # 変数の初期化
        self.excel_dir = ""
        self.excel_files = []
        self.selected_file = tk.StringVar()
        self.selected_region = tk.StringVar(value="並木")  # デフォルト値
        self.regions = []
        self.current_excel_data = None

        # GUIの作成
        self.create_widgets()

    def create_widgets(self):
        # Excelディレクトリ選択
        tk.Button(
            self.root,
            text="Select Excel Directory",
            command=self.select_excel_directory
        ).pack(pady=10)

        # Excelファイル選択
        self.excel_dropdown = ttk.Combobox(
            self.root,
            textvariable=self.selected_file,
            state="readonly"
        )
        self.excel_dropdown.pack(pady=5)

        tk.Button(
            self.root,
            text="Select Excel",
            command=self.load_excel_file
        ).pack(pady=5)

        # 地区選択
        self.region_dropdown = ttk.Combobox(
            self.root,
            textvariable=self.selected_region,
            state="readonly"
        )
        self.region_dropdown.pack(pady=5)

        tk.Button(
            self.root,
            text="Select Region",
            command=self.select_region
        ).pack(pady=5)

        # iCal作成
        tk.Button(
            self.root,
            text="Create iCal",
            command=self.create_ical
        ).pack(pady=10)

    def select_excel_directory(self):
        self.excel_dir = filedialog.askdirectory()
        if self.excel_dir:
            self.excel_files = [
                f for f in os.listdir(self.excel_dir)
                if f.endswith('_calendar.xlsx') and not f.startswith('~$')
            ]
            self.excel_dropdown['values'] = self.excel_files
            if self.excel_files:
                self.selected_file.set(self.excel_files[0])

    def load_excel_file(self):
        if not self.selected_file.get():
            messagebox.showerror("エラー", "Excelファイルを選択してください")
            return

        file_path = os.path.join(self.excel_dir, self.selected_file.get())
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            self.current_excel_data = ws

            # 地区名の取得（2行目以降の1列目）
            self.regions = []
            for row in ws.iter_rows(min_row=2):
                region_value = row[0].value
                if region_value:
                    # 改行を含む場合は最初の行のみを使用
                    region_name = str(region_value).split('\n')[0].strip()
                    if region_name and region_name not in self.regions:
                        self.regions.append(region_name)

            self.region_dropdown['values'] = self.regions

            if "並木" in self.regions:
                self.selected_region.set("並木")
            elif self.regions:
                self.selected_region.set(self.regions[0])

            logger.info(f"読み込んだ地区: {self.regions}")
            messagebox.showinfo("成功", "Excelファイルを読み込みました")

        except Exception as e:
            logger.error(f"Excelファイル読み込みエラー: {str(e)}")
            messagebox.showerror("エラー", f"Excelファイルの読み込みに失敗しました: {str(e)}")

    def select_region(self):
        if not self.selected_region.get():
            messagebox.showerror("エラー", "地区を選択してください")
            return
        messagebox.showinfo("成功", f"地区「{self.selected_region.get()}」を選択しました")

    def create_ical(self):
        if not self.current_excel_data or not self.selected_region.get():
            messagebox.showerror("エラー", "Excelファイルと地区を選択してください")
            return

        try:
            # iCalディレクトリの作成
            os.makedirs('ical', exist_ok=True)

            # 選択された地区の行を検索
            region_row = None
            selected_region = self.selected_region.get()
            for row in self.current_excel_data.iter_rows(min_row=2):
                cell_value = str(row[0].value).split('\n')[0].strip() if row[0].value else ""
                if cell_value == selected_region:
                    region_row = row
                    break

            if not region_row:
                logger.error(f"地区が見つかりません: {selected_region}")
                messagebox.showerror("エラー", "選択された地区が見つかりません")
                return

            # カレンダーの作成
            cal = Calendar()
            cal.add('prodid', '-//Tsukuba Gomi Calendar//JP')
            cal.add('version', '2.0')
            cal.add('calscale', 'GREGORIAN')
            cal.add('method', 'PUBLISH')
            cal.add('x-wr-calname', f'つくば市 {selected_region} ごみ収集カレンダー')
            cal.add('x-wr-timezone', 'Asia/Tokyo')

            # ごみの種類と対応する列のインデックス（0から始まるインデックスに合わせて調整）
            gomi_types = {
                "燃やせるごみ": 2,  # 3列目（インデックスは2）
                "びん": 3,  # 4列目（インデックスは3）
                "スプレー容器": 4,  # 5列目（インデックスは4）
                "ペットボトル": 5,  # 6列目（インデックスは5）
                "燃やせないごみ": 6,  # 7列目（インデックスは6）
                "古紙・古布": 7,  # 8列目（インデックスは7）
                "プラスチック製容器包装": 8,  # 9列目（インデックスは8）
                "かん": 9,  # 10列目（インデックスは9）
                "粗大ごみ（予約制）": 10  # 11列目（インデックスは10）
            }

            # 現在時刻（タイムスタンプ用）
            now = datetime.now(self.timezone)

            # 各ごみ種類の収集日をイベントとして追加
            for gomi_type, col_idx in gomi_types.items():
                # col_idxはそのまま使用（-1しない）
                cell_value = region_row[col_idx].value
                if not cell_value:
                    continue

                dates_str = str(cell_value)
                for date_str in dates_str.split(','):
                    date_str = date_str.strip()
                    if not date_str:
                        continue

                    try:
                        # 日付形式のバリデーション
                        if not all(c.isdigit() or c == '/' for c in date_str):
                            logger.warning(f"不正な日付形式: {date_str}")
                            continue

                        collection_date = datetime.strptime(date_str, '%Y/%m/%d')
                        next_day = collection_date + timedelta(days=1)
                        
                        # イベントの作成
                        event = Event()
                        event.add('summary', f"{gomi_type}の日")
                        event.add('description', f"つくば市 {selected_region} {gomi_type}の収集日です")
                        
                        # 終日イベントとして設定（終了日は次の日を指定）
                        event.add('dtstart', collection_date.date())
                        event.add('dtend', next_day.date())
                        event.add('transp', 'TRANSPARENT')
                        
                        # メタデータの追加
                        event.add('dtstamp', now)
                        event['uid'] = str(uuid.uuid4())
                        
                        cal.add_component(event)
                        logger.info(f"イベント追加: {gomi_type} - {date_str}")

                    except ValueError as e:
                        logger.error(f"日付の解析エラー: {date_str} - {str(e)}")
                        continue

            # ファイル名の作成（YYYYMM_地区名.ics）
            filename = f"{self.selected_file.get().split('_')[0]}_{self.selected_region.get()}.ics"
            filepath = os.path.join('ical', filename)

            # iCalファイルの保存
            with open(filepath, 'wb') as f:
                f.write(cal.to_ical())

            logger.info(f"iCalファイル作成成功: {filename}")
            messagebox.showinfo("成功", f"iCalファイルを作成しました: {filename}")

        except Exception as e:
            logger.error(f"iCalファイル作成エラー: {str(e)}")
            messagebox.showerror("エラー", f"iCalファイルの作成に失敗しました: {str(e)}")

def main():
    root = tk.Tk()
    app = GomiCalendarApp(root)
    root.mainloop()

if __name__ == "__main__":
    # ログ設定
    import logging
    logging.basicConfig(level=logging.INFO)
    logger = logging.getLogger(__name__)
    
    main()